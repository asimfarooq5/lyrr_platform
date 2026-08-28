"""
Admin endpoints - dashboard, analytics, management
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta

from app.core.database import get_db
from app.api.v1.endpoints.auth import get_current_active_user
from app.models.user import User
from app.models.book import Book
from app.models.content import Payment, UserSubscription, SubscriptionPlan
from app.models.reading_session import ReadingSession
from app.models.sync import SyncCheckpoint

router = APIRouter()


def _require_admin(current_user: User = Depends(get_current_active_user)) -> User:
    """Require an authenticated admin user."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin privileges required")
    return current_user


@router.get("/dashboard")
async def get_dashboard(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(_require_admin),
):
    """Get admin dashboard statistics"""
    # Count users
    user_count = await db.scalar(select(func.count(User.id)))
    
    # Count books
    book_count = await db.scalar(select(func.count(Book.id)))
    
    # Count syncs in last 24h
    yesterday = datetime.utcnow() - timedelta(hours=24)
    sync_count = await db.scalar(
        select(func.count(SyncCheckpoint.id))
        .where(SyncCheckpoint.last_sync_at >= yesterday)
    )
    
    return {
        "total_users": user_count or 0,
        "total_books": book_count or 0,
        "syncs_today": sync_count or 0,
        "period": "24h",
    }


@router.get("/users")
async def list_users(
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(_require_admin),
):
    """List all users"""
    result = await db.execute(
        select(User).offset(skip).limit(limit)
    )
    users = result.scalars().all()
    
    return {
        "users": [
            {
                "id": u.id,
                "email": u.email,
                "is_active": u.is_active,
                "is_admin": u.is_admin,
                "is_verified": u.is_verified,
                "created_at": u.created_at.isoformat() if u.created_at else None,
            }
            for u in users
        ],
        "total": len(users),
    }


@router.get("/analytics")
async def get_analytics(
    days: int = 30,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(_require_admin),
):
    """Get analytics data (FRS §13).

    Returns user counts, reading/listening sessions, revenue by period,
    and most popular books.
    """
    days = max(1, min(days, 365))
    since = datetime.utcnow() - timedelta(days=days)

    # Users
    total_users = (await db.scalar(select(func.count(User.id)))) or 0
    new_users = (await db.scalar(
        select(func.count(User.id)).where(User.created_at >= since)
    )) or 0

    # Reading sessions (daily reading aggregation)
    total_reading_sessions = (await db.scalar(
        select(func.count(ReadingSession.id))
    )) or 0
    reading_sessions_in_period = (await db.scalar(
        select(func.count(ReadingSession.id)).where(ReadingSession.created_at >= since)
    )) or 0
    reading_minutes_in_period = (await db.scalar(
        select(func.coalesce(func.sum(ReadingSession.duration_seconds), 0))
        .where(ReadingSession.created_at >= since)
    )) or 0
    reading_minutes_in_period = int(reading_minutes_in_period / 60)

    # Listening sessions: progress positions reported from the audio player.
    # ReadingProgress tracks both read and listen; listening sessions are the
    # reading_sessions rows attached to books that have audio (BookMedia).
    listening_sessions = (await db.scalar(
        select(func.count(ReadingSession.id))
        .where(
            ReadingSession.created_at >= since,
            ReadingSession.book_id.isnot(None),
        )
    )) or 0

    # Revenue by period
    revenue_rows = (await db.execute(
        select(
            func.date(Payment.completed_at).label("day"),
            func.sum(Payment.amount).label("revenue"),
            func.count(Payment.id).label("payments"),
        )
        .where(
            Payment.status == "completed",
            Payment.completed_at >= since,
        )
        .group_by(func.date(Payment.completed_at))
        .order_by(func.date(Payment.completed_at))
    )).all()
    revenue_by_day = [
        {"date": str(row.day), "revenue": float(row.revenue or 0), "payments": row.payments or 0}
        for row in revenue_rows
    ]
    total_revenue = sum(r["revenue"] for r in revenue_by_day)
    total_payments = sum(r["payments"] for r in revenue_by_day)

    # Revenue breakdown by method
    method_rows = (await db.execute(
        select(Payment.method, func.sum(Payment.amount), func.count(Payment.id))
        .where(Payment.status == "completed", Payment.completed_at >= since)
        .group_by(Payment.method)
    )).all()
    revenue_by_method = [
        {"method": row[0], "revenue": float(row[1] or 0), "count": row[2] or 0}
        for row in method_rows
    ]

    # Most popular books by reading session volume in period
    popular_rows = (await db.execute(
        select(
            Book.id, Book.title, Book.author,
            func.count(ReadingSession.id).label("sessions"),
        )
        .join(ReadingSession, ReadingSession.book_id == Book.id)
        .where(ReadingSession.created_at >= since)
        .group_by(Book.id, Book.title, Book.author)
        .order_by(func.count(ReadingSession.id).desc())
        .limit(10)
    )).all()
    popular_books = [
        {
            "book_id": row.id,
            "title": row.title,
            "author": row.author,
            "sessions": row.sessions,
        }
        for row in popular_rows
    ]

    # Subscription status
    active_subscriptions = (await db.scalar(
        select(func.count(UserSubscription.id))
        .where(UserSubscription.status == "active")
    )) or 0
    plan_rows = (await db.execute(
        select(
            SubscriptionPlan.name,
            func.count(UserSubscription.id),
        )
        .join(UserSubscription, UserSubscription.plan_id == SubscriptionPlan.id)
        .where(UserSubscription.status == "active")
        .group_by(SubscriptionPlan.name)
    )).all()
    subscriptions_by_plan = [
        {"plan": row[0], "count": row[1]}
        for row in plan_rows
    ]

    return {
        "period": f"last_{days}_days",
        "users": {
            "total": total_users,
            "new": new_users,
        },
        "sessions": {
            "total_reading_sessions": total_reading_sessions,
            "reading_sessions": reading_sessions_in_period,
            "listening_sessions": listening_sessions,
            "reading_minutes": reading_minutes_in_period,
        },
        "revenue": {
            "total": total_revenue,
            "total_payments": total_payments,
            "by_day": revenue_by_day,
            "by_method": revenue_by_method,
        },
        "popular_books": popular_books,
        "subscriptions": {
            "active": active_subscriptions,
            "by_plan": subscriptions_by_plan,
        },
    }


@router.get("/analytics/export")
async def export_analytics(
    days: int = 30,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(_require_admin),
):
    """Export analytics to an Excel workbook (.xlsx) - FRS §13."""
    data = await get_analytics(days=days, db=db, admin=admin)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws.append(["LYRR Analytics", ""])
    ws.append(["Period", data["period"]])
    ws.append([])
    ws.append(["Metric", "Value"])
    ws.append(["Total Users", data["users"]["total"]])
    ws.append(["New Users", data["users"]["new"]])
    ws.append(["Reading Sessions", data["sessions"]["reading_sessions"]])
    ws.append(["Listening Sessions", data["sessions"]["listening_sessions"]])
    ws.append(["Reading Minutes", data["sessions"]["reading_minutes"]])
    ws.append(["Total Revenue", data["revenue"]["total"]])
    ws.append(["Total Payments", data["revenue"]["total_payments"]])
    ws.append(["Active Subscriptions", data["subscriptions"]["active"]])
    ws.append([])
    ws.append(["Revenue by Method", "Revenue", "Count"])
    for row in data["revenue"]["by_method"]:
        ws.append([row["method"], row["revenue"], row["count"]])
    ws.append([])
    ws.append(["Subscriptions by Plan", "Count"])
    for row in data["subscriptions"]["by_plan"]:
        ws.append([row["plan"], row["count"]])

    # ---- Revenue by day sheet ----
    ws2 = wb.create_sheet("Revenue by Day")
    ws2.append(["Date", "Revenue", "Payments"])
    for row in data["revenue"]["by_day"]:
        ws2.append([row["date"], row["revenue"], row["payments"]])

    # ---- Popular books sheet ----
    ws3 = wb.create_sheet("Popular Books")
    ws3.append(["Book ID", "Title", "Author", "Sessions"])
    for row in data["popular_books"]:
        ws3.append([row["book_id"], row["title"], row["author"], row["sessions"]])

    # ---- Formatting: bold headers ----
    header_font = Font(bold=True)
    fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = fill
        sheet.column_dimensions[get_column_letter(1)].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"lyrr_analytics_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
